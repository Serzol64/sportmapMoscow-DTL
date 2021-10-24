#! /usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json
import geojson
import uuid
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware


def datasetConnect(sheet):

	if sheet == 'objects': return load_workbook('data/db.xlsx').get_sheet_by_name('Объекты отдельно')
	elif sheet == 'access': return load_workbook('data/db.xlsx').get_sheet_by_name('Обозначения доступности')
	elif sheet == 'regional': return load_workbook('data/db.xlsx').get_sheet_by_name('Метагеоданные')
	elif sheet == 'territory': return load_workbook('data/db.xlsx').get_sheet_by_name('Метагеоданные1')
	else: return load_workbook('data/db.xlsx').get_sheet_by_name('Со спортзонами и видами спорта')

app = FastAPI()

origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get('/')
def welcome():
	#Основной сервис для карты

	objectsDB = datasetConnect('objects')

	objId = []
	lat = []
	longitude = []
	geodata = []

	for index in range(len(objectsDB['I'])):
		if objectsDB['H'][index].value != "Долгота (Longitude)":
			objId.append(objectsDB['A'][index].value)
			longitude.append(objectsDB['I'][index].value)
	for index in range(len(objectsDB['H'])):
		if objectsDB['H'][index].value != "Широта (Latitude)":
			lat.append(objectsDB['H'][index].value)


	geodata.append({
		'features': []
	})

	for i in range(len(lat)):
		geodata['features'].append({
			'type': "Feature",
			'geometry': {
				'type': "Point",
				'coordinates': [float(longitude[i]), float(lat[i])]
			},
			'properties': {
				'id': objId[i]
			}
		})

	response = geojson.FeatureCollection(geodata)

	return response


@app.post('/objectInfo')
def objectData(objectId: int):

	response_object = ""
	response_address = ""
	response_zone = ""
	response_zoneType = ""
	response_sport = ""
	response_director = ""


	objectsDB = datasetConnect('objects')
	detalis = datasetConnect('data')


	for index in range(len(detalis['B'])):

		if detalis['B'][index].value != "Объект":
			if objectsDB['A'][index].value == objectId:

				dirState = ""

				if objectsDB['E'][index].value == "NULL": dirState = "Неизвестный владелец"
				else: dirState = detalis['E'][index].value

				response_object = detalis['B'][index].value
				response_address = detalis['C'][index].value
				response_region = detalis['J'][index].value
				response_zone = detalis['G'][index].value
				response_zoneType = detalis['H'][index].value
				response_sport = detalis['K'][index].value
				response_director = dirState




	return {
		'object': response_object,
		'location': response_address,
		'region': response_region,
		'sportzone': response_zone,
		'sportzoneType': response_zoneType,
		'sportType': response_sport,
		'director': response_director
	}








